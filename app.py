# Copyright 2018 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License")
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# [START gae_python37_app]
import os
import io
import functools
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import pandas as pd
import csv
import base64
import json
import traceback
import daff
import pyhornedowl
import networkx
import re

from flask import Flask, request, g, session, redirect, url_for, render_template
from flask import render_template_string, jsonify, Response, send_file
from flask_github import GitHub
from flask_cors import CORS #enable cross origin request?
from flask_caching import Cache

from io import StringIO  #for download
import requests #for download

from sqlalchemy import create_engine, Column, Integer, String
from sqlalchemy.orm import scoped_session, sessionmaker
from sqlalchemy.ext.declarative import declarative_base

from datetime import date

from urllib.request import urlopen

import threading

import whoosh
from whoosh.qparser import MultifieldParser,QueryParser

from datetime import datetime

# setup sqlalchemy

from config import *

engine = create_engine(DATABASE_URI)
db_session = scoped_session(sessionmaker(autocommit=False,
                                         autoflush=False,
                                         bind=engine))
Base = declarative_base()

class User(Base):
    __tablename__ = 'users'

    id = Column(Integer, primary_key=True)
    github_access_token = Column(String(255))
    github_id = Column(Integer)
    github_login = Column(String(255))

    def __init__(self, github_access_token):
        self.github_access_token = github_access_token

class NextId(Base):
    __tablename__ = 'nextids'
    id = Column(Integer,primary_key=True)
    repo_name = Column(String(50))
    next_id = Column(Integer)

def init_db():

    Base.query = db_session.query_property()
    Base.metadata.create_all(bind=engine)


# Create an app instance
class FlaskApp(Flask):
    def __init__(self, *args, **kwargs):
        super(FlaskApp, self).__init__(*args, **kwargs)
        self._activate_background_job()

    def _activate_background_job(self):
        init_db()

# If `entrypoint` is not defined in app.yaml, App Engine will look for an app
# called `app` in `main.py`.
app = FlaskApp(__name__)
CORS(app) #cross origin across all 
app.config['CORS_HEADERS'] = 'Content-Type'
cors = CORS(app, resources={
    r"/api/*":{
        "origins":"*"
    }
})


app.config.from_object('config')
cache = Cache(app) #caching
# set cache for each prefix in prefixes:    
for prefix in PREFIXES: 
    cache.set("latestID" + prefix[0], 0)
# cache.set("latestID",0) #initialise caching
print("cache initialised")

github = GitHub(app)


# Implementation of Google Cloud Storage for index
class BucketStorage(whoosh.filedb.filestore.RamStorage):
    def __init__(self, bucket):
        super().__init__()
        self.bucket = bucket
        self.filenameslist = []

    def save_to_bucket(self):
        for name in self.files.keys():
            with self.open_file(name) as source:
                print("Saving file",name)
                blob = self.bucket.blob(name)
                blob.upload_from_file(source)
        for name in self.filenameslist:
            if name not in self.files.keys():
                blob = self.bucket.blob(name)
                print("Deleting old file",name)
                self.bucket.delete_blob(blob.name)
                self.filenameslist.remove(name)

    def open_from_bucket(self):
        self.filenameslist = []
        for blob in bucket.list_blobs():
            print("Opening blob",blob.name)
            self.filenameslist.append(blob.name)
            f = self.create_file(blob.name)
            blob.download_to_file(f)
            f.close()


class SpreadsheetSearcher:
    # bucket is defined in config.py
    def __init__(self):
        self.storage = BucketStorage(bucket)
        self.threadLock = threading.Lock()

    def searchFor(self, repo_name, search_string="", assigned_user=""):
        self.storage.open_from_bucket()
        ix = self.storage.open_index()

        mparser = MultifieldParser(["class_id","label","definition","parent","tobereviewedby"],
                                schema=ix.schema)

        query = mparser.parse("repo:"+repo_name+
                              (" AND ("+search_string+")" if search_string  else "")+
                              (" AND tobereviewedby:"+assigned_user if assigned_user else "") )

        with ix.searcher() as searcher:
            results = searcher.search(query, limit=100)
            resultslist = []
            for hit in results:
                allfields = {}
                for field in hit:
                    allfields[field]=hit[field]
                resultslist.append(allfields)

        ix.close()
        return (resultslist)

    def updateIndex(self, repo_name, folder, sheet_name, header, sheet_data):
        self.threadLock.acquire()
        print("Updating index...")
        self.storage.open_from_bucket()
        ix = self.storage.open_index()
        writer = ix.writer()
        mparser = MultifieldParser(["repo", "spreadsheet"],
                                   schema=ix.schema)
        print("About to delete for query string: ","repo:" + repo_name + " AND spreadsheet:'" + folder+"/"+sheet_name+"'")
        writer.delete_by_query(
            mparser.parse("repo:" + repo_name + " AND spreadsheet:\"" + folder+"/"+sheet_name+"\""))
        writer.commit()

        writer = ix.writer()

        for r in range(len(sheet_data)):
            row = [v for v in sheet_data[r].values()]
            del row[0] # Tabulator-added ID column

            if "ID" in header:
                class_id = row[header.index("ID")]
            else:
                class_id = None
            if "Label" in header:
                label = row[header.index("Label")]
            else:
                label = None
            if "Definition" in header:
                definition = row[header.index("Definition")]
            else:
                definition = None
            if "Parent" in header:
                parent = row[header.index("Parent")]
            else:
                parent = None
            if "To be reviewed by" in header:
                tobereviewedby = row[header.index("To be reviewed by")]
            else:
                tobereviewedby = None

            if class_id or label or definition or parent:
                writer.add_document(repo=repo_name,
                                    spreadsheet=folder+'/'+sheet_name,
                                    class_id=(class_id if class_id else None),
                                    label=(label if label else None),
                                    definition=(definition if definition else None),
                                    parent=(parent if parent else None),
                                    tobereviewedby=(tobereviewedby if tobereviewedby else None))
        writer.commit(optimize=True)
        self.storage.save_to_bucket()
        ix.close()
        self.threadLock.release()
        print("Update of index completed.")

    def getNextId(self,repo_name):
        self.threadLock.acquire()
        self.storage.open_from_bucket()
        ix = self.storage.open_index()

        nextId = 0

        mparser = QueryParser("class_id",
                              schema=ix.schema)
        if repo_name == "BCIO":
            updated_repo_name ="BCIO:" # in order to eliminate "BCIOR" from results
        else:
            updated_repo_name = repo_name
        query = mparser.parse(updated_repo_name.upper()+"*")
        # print("searching ", repo_name)
        with ix.searcher() as searcher:
            results = searcher.search(query, sortedby="class_id",reverse=True)
            tophit = results[0]
            mostRecentID = cache.get("latestID"+repo_name) # check latest ID 
            if mostRecentID is None: # error check no cache set
                mostRecentID = 0
                print("error latestID",repo_name," was None!")
                cache.set("latestID"+repo_name, 0)
            nextId = int(tophit['class_id'].split(":")[1] )+1

            # check nextId against cached most recent id:
            if not(nextId > mostRecentID):
                print("cached version is higher: ", mostRecentID, " > ", nextId)
                nextId = cache.get("latestID"+repo_name)+1                
            cache.set("latestID"+repo_name, nextId)
            

        ix.close()

        self.threadLock.release()
        return (nextId)


searcher = SpreadsheetSearcher()

class OntologyDataStore:
    node_props = {"shape":"box","style":"rounded", "font": "helvetica"}
    rel_cols = {"has part":"blue","part of":"blue","contains":"green",
                "has role":"darkgreen","is about":"darkgrey",
                "has participant":"darkblue"}

    def __init__(self):
        self.releases = {}
        self.releasedates = {}
        self.label_to_id = {}
        self.graphs = {}

    def parseRelease(self,repo):
        # Keep track of when you parsed this release
        self.graphs[repo] = networkx.MultiDiGraph()
        self.releasedates[repo] = date.today()
        #print("Release date ",self.releasedates[repo])

        # Get the ontology from the repository
        ontofilename = app.config['RELEASE_FILES'][repo]
        repositories = app.config['REPOSITORIES']
        repo_detail = repositories[repo]
        location = f"https://raw.githubusercontent.com/{repo_detail}/master/{ontofilename}"
        print("Fetching release file from", location)
        data = urlopen(location).read()  # bytes
        ontofile = data.decode('utf-8')

        # Parse it
        if ontofile:
            self.releases[repo] = pyhornedowl.open_ontology(ontofile)
            prefixes = app.config['PREFIXES']
            for prefix in prefixes:
                self.releases[repo].add_prefix_mapping(prefix[0],prefix[1])
            for classIri in self.releases[repo].get_classes():
                classId = self.releases[repo].get_id_for_iri(classIri)
                if classId:
                    classId = classId.replace(":","_")
                    # is it already in the graph?
                    if classId not in self.graphs[repo].nodes:
                        label = self.releases[repo].get_annotation(classIri, app.config['RDFSLABEL'])
                        if label:
                            self.label_to_id[label.strip()] = classId
                            self.graphs[repo].add_node(classId,
                                                       label=label.strip().replace(" ", "\n"),
                                                **OntologyDataStore.node_props)
                        else:
                            print("Could not determine label for IRI",classIri)
                else:
                    print("Could not determine ID for IRI",classIri)
            for classIri in self.releases[repo].get_classes():
                classId = self.releases[repo].get_id_for_iri(classIri)
                if classId:
                    parents = self.releases[repo].get_superclasses(classIri)
                    for p in parents:
                        plabel = self.releases[repo].get_annotation(p, app.config['RDFSLABEL'])
                        if plabel and plabel.strip() in self.label_to_id:
                            self.graphs[repo].add_edge(self.label_to_id[plabel.strip()],
                                                       classId.replace(":", "_"), dir="back")
                    axioms = self.releases[repo].get_axioms_for_iri(classIri) # other relationships
                    for a in axioms:
                        # Example: ['SubClassOf', 'http://purl.obolibrary.org/obo/CHEBI_27732', ['ObjectSomeValuesFrom', 'http://purl.obolibrary.org/obo/RO_0000087', 'http://purl.obolibrary.org/obo/CHEBI_60809']]
                        if len(a) == 3 and a[0]=='SubClassOf' \
                            and isinstance(a[2], list) and len(a[2])==3 \
                            and a[2][0]=='ObjectSomeValuesFrom':
                            relIri = a[2][1]
                            targetIri = a[2][2]
                            rel_name = self.releases[repo].get_annotation(relIri, app.config['RDFSLABEL'])
                            targetLabel = self.releases[repo].get_annotation(targetIri, app.config['RDFSLABEL'])
                            if targetLabel and targetLabel.strip() in self.label_to_id:
                                if rel_name in OntologyDataStore.rel_cols:
                                    rcolour = OntologyDataStore.rel_cols[rel_name]
                                else:
                                    rcolour = "orange"
                                self.graphs[repo].add_edge(classId.replace(":", "_"),
                                                           self.label_to_id[targetLabel.strip()],
                                                           color=rcolour,
                                                           label=rel_name)

    def getReleaseLabels(self, repo):
        all_labels = set()
        for classIri in self.releases[repo].get_classes():
            all_labels.add(self.releases[repo].get_annotation(classIri, app.config['RDFSLABEL']))
        return( all_labels )
    
    #todo: did this work? 
    def getReleaseIDs(self, repo):
        all_IDs = set()
        for classIri in self.releases[repo].get_classes():
            classId = self.releases[repo].get_id_for_iri(classIri)
            all_IDs.add(classId)
        return( all_IDs )

    def parseSheetData(self, repo, data):
        for entry in data:
            if 'ID' in entry and \
                    'Label' in entry and \
                    'Definition' in entry and \
                    'Parent' in entry and \
                    len(entry['ID'])>0:
                entryId = entry['ID'].replace(":", "_")
                entryLabel = entry['Label'].strip()
                self.label_to_id[entryLabel] = entryId
                self.graphs[repo].add_node(entryId, label=entryLabel.replace(" ", "\n"), **OntologyDataStore.node_props)
                if entryId in self.graphs[repo].nodes:
                    self.graphs[repo].remove_node(entryId)
                    self.graphs[repo].add_node(entryId, label=entryLabel.replace(" ", "\n"), **OntologyDataStore.node_props)
        for entry in data:
            if 'ID' in entry and \
                    'Label' in entry and \
                    'Definition' in entry and \
                    'Parent' in entry and \
                    len(entry['ID'])>0:
                entryParent = re.sub("[\[].*?[\]]", "", str(entry['Parent'])).strip()
                if entryParent in self.label_to_id:  # Subclass relations
                    # Subclass relations must be reversed for layout
                    self.graphs[repo].add_edge(self.label_to_id[entryParent],
                                               entry['ID'].replace(":", "_"), dir="back")
                for header in entry.keys():  # Other relations
                    if entry[header] and str(entry[header]).strip() and "REL" in header:
                        # Get the rel name
                        rel_names = re.findall(r"'([^']+)'", header)
                        if len(rel_names) > 0:
                            rel_name = rel_names[0]
                            if rel_name in OntologyDataStore.rel_cols:
                                rcolour = OntologyDataStore.rel_cols[rel_name]
                            else:
                                rcolour = "orange"

                            relValues = entry[header].split(";")
                            for relValue in relValues:
                                if relValue.strip() in self.label_to_id:
                                    self.graphs[repo].add_edge(entry['ID'].replace(":", "_"),
                                                               self.label_to_id[relValue.strip()],
                                                               color=rcolour,
                                                               label=rel_name)
 
 # re-factored the following:  
    # *new : getIDsFromSheetMultiSelect
    # getIDsFromSheet - related ID's from whole sheet
    # getIDsFromSelection - related ID's from selection in sheet    
    # getRelatedIds - related ID's from list of ID's

    # *new: getIDSForSheetGraphMultiSelect
    # getDotForSheetGraph - graph from whole sheet
    # getDotForSelection - graph from selection in sheet
    # getDotForIDs - graph from ID list
 
    def getIDsFromSheetMultiSelect(self, repo, data, filter):
        ids = []
        for entry in data:
            if 'Curation status' in entry and str(entry['Curation status']) == "Obsolete": 
                print("Obsolete: ", entry)
            else:
                if filter != [""] and filter != []:
                    for f in filter:
                        if str(entry['Curation status']) == f:
                            if 'ID' in entry and len(entry['ID'])>0:
                                ids.append(entry['ID'].replace(":","_"))
                            if 'Parent' in entry:
                                entryParent = re.sub("[\[].*?[\]]", "", str(entry['Parent'])).strip()
                                if entryParent in self.label_to_id:
                                    ids.append(self.label_to_id[entryParent])
                            if ":" in entry['ID'] or "_" in entry['ID']:
                                entryIri = self.releases[repo].get_iri_for_id(entry['ID'].replace("_", ":"))                    
                                if entryIri:
                                    descs = pyhornedowl.get_descendants(self.releases[repo], entryIri)
                                    for d in descs:
                                        ids.append(self.releases[repo].get_id_for_iri(d).replace(":", "_"))
                            if self.graphs[repo]:
                                graph_descs = None
                                try:
                                    graph_descs = networkx.algorithms.dag.descendants(self.graphs[repo],entry['ID'].replace(":", "_"))
                                except networkx.exception.NetworkXError:
                                    print("NetworkXError sheet multiselect: ", entry['ID'])
                                
                                if graph_descs is not None:
                                    for g in graph_descs:
                                        if g not in ids:
                                            ids.append(g)
                            
                            
        return (ids)
    
    def getIDsFromSheet(self, repo, data, filter):
        # list of ids from sheetExternal
        ids = []
        for entry in data:
            if 'Curation status' in entry and str(entry['Curation status']) == "Obsolete": 
                print("Obsolete: ", entry)
            else:
                if filter != "":
                    if str(entry['Curation status']) == filter:
                        if 'ID' in entry and len(entry['ID'])>0:
                            ids.append(entry['ID'].replace(":","_"))

                        if 'Parent' in entry:
                            entryParent = re.sub("[\[].*?[\]]", "", str(entry['Parent'])).strip()
                            if entryParent in self.label_to_id:
                                ids.append(self.label_to_id[entryParent])
                        
                        if ":" in entry['ID'] or "_" in entry['ID']:
                            entryIri = self.releases[repo].get_iri_for_id(entry['ID'].replace("_", ":"))                    
                            if entryIri:
                                descs = pyhornedowl.get_descendants(self.releases[repo], entryIri)
                                for d in descs:
                                    ids.append(self.releases[repo].get_id_for_iri(d).replace(":", "_"))
                        if self.graphs[repo]:
                            graph_descs = None
                            try:
                                graph_descs = networkx.algorithms.dag.descendants(self.graphs[repo],entry['ID'].replace(":", "_"))
                            except networkx.exception.NetworkXError:
                                print("NetworkXError sheet filter: ", entry['ID'])
                            
                            if graph_descs is not None:
                                for g in graph_descs:
                                    if g not in ids:
                                        ids.append(g)  
                else:
                    if 'ID' in entry and len(entry['ID'])>0:
                            ids.append(entry['ID'].replace(":","_"))

                    if 'Parent' in entry:
                        entryParent = re.sub("[\[].*?[\]]", "", str(entry['Parent'])).strip()
                        print("found entryParent: ", entryParent)
                        if entryParent in self.label_to_id:
                            ids.append(self.label_to_id[entryParent])
                    if ":" in entry['ID'] or "_" in entry['ID']: 
                        entryIri = self.releases[repo].get_iri_for_id(entry['ID'].replace("_", ":"))                    
                        if entryIri:
                            descs = pyhornedowl.get_descendants(self.releases[repo], entryIri)
                            for d in descs:
                                ids.append(self.releases[repo].get_id_for_iri(d).replace(":", "_"))
                    if self.graphs[repo]:
                        graph_descs = None
                        try:
                            graph_descs = networkx.algorithms.dag.descendants(self.graphs[repo],entry['ID'].replace(":", "_"))
                        except networkx.exception.NetworkXError:
                            print("NetworkXError Sheet: ", entry['ID'])
                        
                        if graph_descs is not None:
                            for g in graph_descs:
                                if g not in ids:
                                    ids.append(g)   
        return (ids)
    
    def getIDsFromSelectionMultiSelect(self, repo, data, selectedIds, filter):
        # Add all descendents of the selected IDs, the IDs and their parents.
        ids = [] 
        for id in selectedIds:
            entry = data[id]
            # don't visualise rows which are set to "Obsolete":
            if 'Curation status' in entry and str(entry['Curation status']) == "Obsolete": 
                pass
            else:
                if filter != [""] and filter != []:
                    for f in filter:
                        if str(entry['Curation status']) == f:
                            if str(entry['ID']) and str(entry['ID']).strip(): #check for none and blank ID's
                                if 'ID' in entry and len(entry['ID']) > 0:
                                    ids.append(entry['ID'].replace(":", "_"))
                                if 'Parent' in entry:
                                    entryParent = re.sub("[\[].*?[\]]", "", str(entry['Parent'])).strip()
                                    if entryParent in self.label_to_id:
                                        ids.append(self.label_to_id[entryParent])
                                if ":" in entry['ID'] or "_" in entry['ID']:
                                    entryIri = self.releases[repo].get_iri_for_id(entry['ID'].replace("_", ":"))
                                    if entryIri:
                                        descs = pyhornedowl.get_descendants(self.releases[repo], entryIri)
                                        for d in descs:
                                            ids.append(self.releases[repo].get_id_for_iri(d).replace(":", "_"))
                                if self.graphs[repo]:
                                    graph_descs = None
                                    try:
                                        graph_descs = networkx.algorithms.dag.descendants(self.graphs[repo],entry['ID'].replace(":", "_"))
                                    except networkx.exception.NetworkXError:
                                        print("NetworkXError selection multiselect: ", entry['ID'])
                                    
                                    if graph_descs is not None:
                                        for g in graph_descs:
                                            if g not in ids:
                                                ids.append(g)                        
        return (ids)

    def getIDsFromSelection(self, repo, data, selectedIds, filter):
        # Add all descendents of the selected IDs, the IDs and their parents.
        ids = [] 
        for id in selectedIds:
            entry = data[id]
            # don't visualise rows which are set to "Obsolete":
            if 'Curation status' in entry and str(entry['Curation status']) == "Obsolete": 
                pass
            else:
                if filter != "":
                    if str(entry['Curation status']) == filter:
                        if str(entry['ID']) and str(entry['ID']).strip(): #check for none and blank ID's
                            if 'ID' in entry and len(entry['ID']) > 0:
                                ids.append(entry['ID'].replace(":", "_"))
                            if 'Parent' in entry:
                                entryParent = re.sub("[\[].*?[\]]", "", str(entry['Parent'])).strip()                                
                                if entryParent in self.label_to_id:
                                    ids.append(self.label_to_id[entryParent])
                            if ":" in entry['ID'] or "_" in entry['ID']:
                                entryIri = self.releases[repo].get_iri_for_id(entry['ID'].replace("_", ":"))
                                if entryIri:
                                    descs = pyhornedowl.get_descendants(self.releases[repo], entryIri)
                                    for d in descs:
                                        ids.append(self.releases[repo].get_id_for_iri(d).replace(":", "_"))
                            if self.graphs[repo]:
                                graph_descs = None
                                try:
                                    graph_descs = networkx.algorithms.dag.descendants(self.graphs[repo],entry['ID'].replace(":", "_"))
                                except networkx.exception.NetworkXError:
                                    print("NetworkXError selection filter: ", str(entry['ID']))
                                
                                if graph_descs is not None:
                                    for g in graph_descs:
                                        if g not in ids:
                                            ids.append(g)  
                else:
                    if str(entry['ID']) and str(entry['ID']).strip(): #check for none and blank ID's
                        if 'ID' in entry and len(entry['ID']) > 0:
                            ids.append(entry['ID'].replace(":", "_"))                            
                        if 'Parent' in entry:
                            entryParent = re.sub("[\[].*?[\]]", "", str(entry['Parent'])).strip() 
                            if entryParent in self.label_to_id:
                                    ids.append(self.label_to_id[entryParent])
                        if ":" in entry['ID'] or "_" in entry['ID']:
                            entryIri = self.releases[repo].get_iri_for_id(entry['ID'].replace("_", ":"))
                            if entryIri:
                                descs = pyhornedowl.get_descendants(self.releases[repo], entryIri)
                                for d in descs:
                                    ids.append(self.releases[repo].get_id_for_iri(d).replace(":", "_"))
                        if self.graphs[repo]:
                            graph_descs = None
                            try:
                                graph_descs = networkx.algorithms.dag.descendants(self.graphs[repo],entry['ID'].replace(":", "_"))
                            except networkx.exception.NetworkXError:                               
                                print("NetworkXError selection all: ", str(entry['ID']))
                            
                            if graph_descs is not None:
                                for g in graph_descs:
                                    if g not in ids:
                                        ids.append(g)                      
        return (ids)

    def getRelatedIDs(self, repo, selectedIds):
        # Add all descendents of the selected IDs, the IDs and their parents.
        # print(selectedIds)
        ids = []
        for id in selectedIds:
            try: 
                # print("got one", id)
                ids.append(id.replace(":","_"))
                if ":" in id or "_" in id: 
                    entryIri = self.releases[repo].get_iri_for_id(id.replace("_", ":"))

                    if entryIri:
                        descs = pyhornedowl.get_descendants(self.releases[repo],entryIri)
                        for d in descs:
                            ids.append(self.releases[repo].get_id_for_iri(d).replace(":","_"))
                            
                        superclasses = self.releases[repo].get_superclasses(entryIri)
                        for s in superclasses:
                            ids.append(self.releases[repo].get_id_for_iri(s).replace(":", "_"))
                if self.graphs[repo]:
                    graph_descs = None
                    try:
                        graph_descs = networkx.algorithms.dag.descendants(self.graphs[repo], id.replace(":", "_"))
                    except networkx.exception.NetworkXError:
                        print("NetworkXError relatedIDs: ", str(id))                    

                    if graph_descs is not None:
                        for g in graph_descs:
                            if g not in ids:
                                ids.append(g)
            except:
                pass
        return (ids)

    def getDotForSheetGraph(self, repo, data, filter):
        # Get a list of IDs from the sheet graph
        #todo: is there a better way to do this? 
        if hasattr(filter, 'lower'): #check if filter is a string
            ids = OntologyDataStore.getIDsFromSheet(self, repo, data, filter)
        else: #should be a list then
            ids = OntologyDataStore.getIDsFromSheetMultiSelect(self, repo, data, filter)             
        subgraph = self.graphs[repo].subgraph(ids)
        P = networkx.nx_pydot.to_pydot(subgraph)
        return (P)

    def getDotForSelection(self, repo, data, selectedIds, filter):
        # Add all descendents of the selected IDs, the IDs and their parents.
        #todo: is there a better way to do this? 
        if hasattr(filter, 'lower'): #check if filter is a string
            ids = OntologyDataStore.getIDsFromSelection(self, repo, data, selectedIds, filter)
        else: #should be a list then
            ids = OntologyDataStore.getIDsFromSelectionMultiSelect(self, repo, data, selectedIds, filter)
        # Then get the subgraph as usual
        subgraph = self.graphs[repo].subgraph(ids)
        P = networkx.nx_pydot.to_pydot(subgraph)
        return (P)

    def getDotForIDs(self, repo, selectedIds):
        # Add all descendents of the selected IDs, the IDs and their parents.
        ids = OntologyDataStore.getRelatedIDs(self, repo, selectedIds)
        # print("got IDS", ids) 
        #todo: is the full repo too many to visualise? 
        #test with part of repo ids:
        # ids = ['ADDICTO_0000678', 'ADDICTO_0000649', 'ADDICTO_0000101', 'ADDICTO_0000101', 'ADDICTO_0000697', 'ADDICTO_0000747', 'ADDICTO_0000747', 'ADDICTO_0000752', 'ADDICTO_0000343', 'ADDICTO_0000343', 'ADDICTO_0000515', 'CHEBI_25693', 'CHEBI_32692', 'CHEBI_119915', 'CHEBI_25693', 'CHEBI_24532', 'CHEBI_5686', 'CHEBI_33595', 'ADDICTO_0001098', 'ADDICTO_0001098', 'ADDICTO_0000431', 'BCIO_036000', 'ADDICTO_0000646', 'ADDICTO_0000708', 'ADDICTO_0000673', 'BCIO_036000', 'ADDICTO_0000687', 'ADDICTO_0000645', 'BCIO_040000', 'OGMS_0000031', 'OGMS_0000031', 'BFO_0000016', 'ADDICTO_0000818', 'ADDICTO_0000818', 'ADDICTO_0000536', 'ADDICTO_0000409', 'ADDICTO_0000409', 'ADDICTO_0000399', 'ADDICTO_0000631']
        # Then get the subgraph as usual
        subgraph = self.graphs[repo].subgraph(ids)
        P = networkx.nx_pydot.to_pydot(subgraph)
        return (P)   

    #to create a dictionary and add all info to it, in the relevant place
    def getMetaData(self, repo, allIDS):                    
        DEFN = "http://purl.obolibrary.org/obo/IAO_0000115"
        SYN = "http://purl.obolibrary.org/obo/IAO_0000118"

        label = "" 
        definition = ""
        synonyms = ""
        entries = []

        
        all_labels = set()
        for classIri in self.releases[repo].get_classes():
            classId = self.releases[repo].get_id_for_iri(classIri).replace(":", "_")
            for id in allIDS:   
                if id is not None:         
                    if classId == id:
                        label = self.releases[repo].get_annotation(classIri, app.config['RDFSLABEL']) #yes
                        iri = self.releases[repo].get_iri_for_label(label)
                        if self.releases[repo].get_annotation(classIri, DEFN) is not None:             
                            definition = self.releases[repo].get_annotation(classIri, DEFN).replace(",", "").replace("'", "").replace("\"", "")   
                        else:
                            definition = ""
                        if self.releases[repo].get_annotation(classIri, SYN) is not None:
                            synonyms = self.releases[repo].get_annotation(classIri, SYN).replace(",", "").replace("'", "").replace("\"", "") 
                        else:
                            synonyms = ""
                        entries.append({
                            "id": id,
                            "label": label, 
                            "synonyms": synonyms,
                            "definition": definition,                      
                        })
        return (entries)



ontodb = OntologyDataStore()

def verify_logged_in(fn):
    """
    Decorator used to make sure that the user is logged in
    """
    @functools.wraps(fn)
    def wrapped(*args, **kwargs):
        # If the user is not logged in, then redirect him to the "logged out" page:
        if not g.user:
            return redirect(url_for("login"))
        return fn(*args, **kwargs)

    return wrapped

# Pages:


@app.before_request
def before_request():
    g.user = None
    if 'user_id' in session:
        #print("user-id in session: ",session['user_id'])
        g.user = User.query.get(session['user_id'])


@app.after_request
def after_request(response):
    db_session.remove()
    # print("after_request is running")
    return response

@app.teardown_request
def teardown_request_func(error=None):
    try:
        db_session.remove()
    except Exception as e:
        print("Error in teardown_request_func: ", str(e))
    # print("teardown_request is running!")
    if error:
        print(str(error))

@github.access_token_getter
def token_getter():
    user = g.user
    if user is not None:
        return user.github_access_token


@app.route('/github-callback')
@github.authorized_handler
def authorized(access_token):
    next_url = request.args.get('next') or url_for('home')
    if access_token is None:
        print("Authorization failed.")
        return redirect(url_for('logout'))

    user = User.query.filter_by(github_access_token=access_token).first()
    if user is None:
        user = User(access_token)
    # Not necessary to get these details here
    # but it helps humans to identify users easily.
    g.user = user
    github_user = github.get('/user')
    user.github_id = github_user['id']
    user.github_login = github_user['login']
    user.github_access_token = access_token
    db_session.add(user)
    db_session.commit()

    session['user_id'] = user.id
    return redirect(next_url)


@app.route('/login')
def login():
    if session.get('user_id', None) is not None:
        session.pop('user_id',None) # Could be stale
    return github.authorize(scope="user,repo")

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('loggedout'))

@app.route("/loggedout")
def loggedout():
    """
    Displays the page to be shown to logged out users.
    """
    return render_template("loggedout.html")


@app.route('/user')
@verify_logged_in
def user():
    return jsonify(github.get('/user'))



# Pages for the app


@app.route('/search', methods=['POST'])
@verify_logged_in
def search():
    searchTerm = request.form.get("inputText")
    repoName = request.form.get("repoName")
    searchResults = searchAcrossSheets(repoName, searchTerm)    
    searchResultsTable = json.dumps(searchResults)
    return ( json.dumps({"message":"Success",
                             "searchResults": searchResultsTable}), 200 )


@app.route('/searchAssignedToMe', methods=['POST'])
@verify_logged_in
def searchAssignedToMe():
    initials = request.form.get("initials")
    print("Searching for initials: " + initials)
    repoName = request.form.get("repoName")
    #below is searching in "Label" column? 
    searchResults = searchAssignedTo(repoName, initials)
    searchResultsTable = json.dumps(searchResults)
    return ( json.dumps({"message":"Success",
                             "searchResults": searchResultsTable}), 200 )
                      

@app.route('/')
@app.route('/home')
@verify_logged_in
def home():
    repositories = app.config['REPOSITORIES']
    user_repos = repositories.keys()
    # Filter just the repositories that the user can see
    if g.user.github_login in USERS_METADATA:
        user_repos = USERS_METADATA[g.user.github_login]["repositories"]

    repositories = {k:v for k,v in repositories.items() if k in user_repos}

    return render_template('index.html',
                           login=g.user.github_login,
                           repos=repositories)


@app.route('/repo/<repo_key>')
@app.route('/repo/<repo_key>/<path:folder_path>')
@verify_logged_in
def repo(repo_key, folder_path=""):
    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]
    directories = github.get(
        f'repos/{repo_detail}/contents/{folder_path}'
    )
    dirs = []
    spreadsheets = []
    #go to edit_external: 
    if folder_path == 'imports': 
        return redirect(url_for('edit_external', repo_key=repo_key, folder_path=folder_path))
    for directory in directories:
        if directory['type']=='dir':
            dirs.append(directory['name'])
        elif directory['type']=='file' and '.xlsx' in directory['name']:
            spreadsheets.append(directory['name'])
    if g.user.github_login in USERS_METADATA:
        user_initials = USERS_METADATA[g.user.github_login]["initials"]
    else:
        print(f"The user {g.user.github_login} has no known metadata")
        user_initials = g.user.github_login[0:2]

    return render_template('repo.html',
                            login=g.user.github_login,
                            user_initials=user_initials,
                            repo_name = repo_key,
                            folder_path = folder_path,
                            directories = dirs,
                            spreadsheets = spreadsheets,
                            )

@app.route("/direct", methods=["POST"])
@verify_logged_in
def direct():
    if request.method == "POST":
        type = json.loads(request.form.get("type"))
        repo = json.loads(request.form.get("repo"))
        sheet = json.loads(request.form.get("sheet"))
        go_to_row = json.loads(request.form.get("go_to_row"))
    repoStr = repo['repo']
    sheetStr = sheet['sheet']
    url = '/edit' + '/' + repoStr + '/' + sheetStr 
    session['type'] = type['type']
    session['label'] = go_to_row['go_to_row']
    session['url'] = url
    return('success')

@app.route("/validate", methods=["POST"]) 
@verify_logged_in
def verify():
    if request.method == "POST":
        cell = json.loads(request.form.get("cell"))
        column = json.loads(request.form.get("column"))
        rowData = json.loads(request.form.get("rowData"))
        headers = json.loads(request.form.get("headers")) 
        table = json.loads(request.form.get("table")) 
    # check for blank cells under conditions first:
    blank = {}
    unique = {}
    returnData, uniqueData = checkBlankMulti(1, blank, unique, cell, column, headers, rowData, table)
    if len(returnData) > 0 or len(uniqueData) > 0:
        return (json.dumps({"message":"fail","values":returnData, "unique":uniqueData}))
    return ('success') 
    

@app.route("/generate", methods=["POST"])
@verify_logged_in
def generate():
    if request.method == "POST":
        repo_key = request.form.get("repo_key")
        rowData = json.loads(request.form.get("rowData"))
        values = {}
        ids = {}
        for row in rowData:
            nextIdStr = str(searcher.getNextId(repo_key))
            fill_num = app.config['DIGIT_COUNT']
            if repo_key == "BCIO":
                fill_num = fill_num-1
            else:
                fill_num = fill_num
            id = repo_key.upper()+":"+nextIdStr.zfill(fill_num)
            ids["ID"+str(row['id'])] = str(row['id'])
            values["ID"+str(row['id'])] = id
        return (json.dumps({"message": "idlist", "IDs": ids, "values": values})) #need to return an array 
    return ('success')  

# validation checks here: 

# recursive check each cell in rowData:
def checkBlankMulti(current, blank, unique, cell, column, headers, rowData, table):
    for index, (key, value) in enumerate(rowData.items()): # todo: really, we need to loop here, surely there is a faster way?
        if index == current:
            if key == "Label" or key == "Definition" or key == "Parent" or key == "Sub-ontology" or key == "Curation status" :
                if key == "Definition" or key == "Parent":
                    status = rowData.get("Curation status") #check for "Curation status"
                    if(status):
                        if rowData["Curation status"] == "Proposed" or rowData["Curation status"] == "External":
                            pass
                        else:
                            if value.strip() == "":
                                blank.update({key:value})
                    else:
                        pass #no "Curation status" column
                else:       
                    if value.strip()=="":
                        blank.update({key:value})
                    else:
                        pass
            if key == "Label" or key == "ID" or key == "Definition":
                if checkNotUnique(value, key, headers, table):
                    unique.update({key:value})
    # go again:
    current = current + 1
    if current >= len(rowData):   
        return (blank, unique)
    return checkBlankMulti(current, blank, unique, cell, column, headers, rowData, table)

def checkNotUnique(cell, column, headers, table):
    counter = 0
    cellStr = cell.strip()
    if cellStr == "":
        return False
    # if Label, ID or Definition column, check cell against all other cells in the same column and return true if same
    for r in range(len(table)): 
        row = [v for v in table[r].values()]
        del row[0] # remove extra numbered "id" column
        for c in range(len(headers)):
            if headers[c] == "ID" and column == "ID":
                if row[c].strip()==cellStr:
                    counter += 1 
                    if counter > 1: #more than one of the same
                        return True
            if headers[c] == "Label" and column == "Label":
                if row[c].strip()==cellStr:
                    counter += 1 
                    if counter > 1: 
                        return True
            if headers[c] == "Definition" and column == "Definition":
                if row[c].strip()==cellStr:
                    counter += 1 
                    if counter > 1: 
                        return True
    return False

@app.route('/edit/<repo_key>/<path:folder>/<spreadsheet>')
@verify_logged_in
def edit(repo_key, folder, spreadsheet):
    if session.get('label') == None:
        go_to_row = ""
    else:
        go_to_row = session.get('label')
        session.pop('label', None)

    if session.get('type') == None:
        type = ""
    else:
        type = session.get('type')
        session.pop('type', None)
    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]
    (file_sha,rows,header) = get_spreadsheet(repo_detail,folder,spreadsheet)
    if g.user.github_login in USERS_METADATA:
        user_initials = USERS_METADATA[g.user.github_login]["initials"]
    else:
        print(f"The user {g.user.github_login} has no known metadata")
        user_initials = g.user.github_login[0:2]
    #Build suggestions data:
    if repo_key not in ontodb.releases or date.today() > ontodb.releasedates[repo_key]:
        ontodb.parseRelease(repo_key)
    suggestions = ontodb.getReleaseLabels(repo_key)
    suggestions = list(dict.fromkeys(suggestions))

    return render_template('edit.html',
                            login=g.user.github_login,
                            user_initials=user_initials,
                            all_initials=ALL_USERS_INITIALS,
                            repo_name = repo_key,
                            folder = folder,
                            spreadsheet_name=spreadsheet,
                            header=json.dumps(header),
                            rows=json.dumps(rows),
                            file_sha = file_sha,
                            go_to_row = go_to_row,
                            type = type, 
                            suggestions = json.dumps(suggestions)
                            )

@app.route('/download_spreadsheet', methods=['POST'])
@verify_logged_in
def download_spreadsheet():
    repo_key = request.form.get("repo_key")
    folder = request.form.get("folder")
    spreadsheet = request.form.get("spreadsheet")
    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]
    url = github.get(f"repos/{repo_detail}/contents/{folder}/{spreadsheet}")
    download_url = url['download_url']
    print(download_url)
    return ( json.dumps({"message":"Success",
            "download_url": download_url}), 200 )
    # return redirect(download_url) #why not?
    # r = requests.get(url)
    # strIO = StringIO.StringIO(r.content)
    # return send_file(strIO, as_attachment=True, attachment_filename={spreadsheet})

    #todo: get spreadsheet location and return it  f"repos/{repo_detail}/contents/{folder}/{spreadsheet}"
    # spreadsheet_file = github.get(f"repos/{repo_detail}/contents/{folder}/{spreadsheet}");
    # spreadsheet_file = github.get(
    #     f'repos/{repo_detail}/contents/{folder}/{spreadsheet}'
    # )
    # base64_bytes = spreadsheet_file['content'].encode('utf-8')
    # decoded_data = base64.decodebytes(base64_bytes)
    # bytesIO = io.BytesIO(decoded_data)
    # wb = openpyxl.load_workbook(io.BytesIO(decoded_data))
    # sheet = wb.active
    # wb.save(spreadsheet)
    # bytesIO.seek(0)  # go to the beginning of the stream
    # #
    # return send_file(
    #     bytesIO,
    #     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    #     attachment_filename=f"{spreadsheet}.xlsx",
    #     as_attachment=True,
    #     cache_timeout=0
    # )
    # return ( json.dumps({"message":"Success",
    #                             "spreadsheet_file": spreadsheet_file}), 200 )
    # return send_file(spreadsheet_file, as_attachment=True, attachment_filename=spreadsheet)
    # return redirect(f"https://raw.githubusercontent.com/{repo_key}/{folder}/{spreadsheet}?token={g.user.github_access_token}")

@app.route('/save', methods=['POST'])
@verify_logged_in
def save():
    repo_key = request.form.get("repo_key")
    folder = request.form.get("folder")
    spreadsheet = request.form.get("spreadsheet")
    row_data = request.form.get("rowData")
    initial_data = request.form.get("initialData")
    file_sha = request.form.get("file_sha").strip()
    commit_msg = request.form.get("commit_msg")
    commit_msg_extra = request.form.get("commit_msg_extra")
    overwrite = False
    overwriteVal = request.form.get("overwrite") 
    if overwriteVal == "true":
        overwrite = True

    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]
    restart = False # for refreshing the sheet (new ID's)
    try:
        initial_data_parsed = json.loads(initial_data)
        row_data_parsed = json.loads(row_data)
        # Get the data, skip the first 'id' column
        initial_first_row = initial_data_parsed[0]
        initial_header = [k for k in initial_first_row.keys()]
        del initial_header[0]
        # Sort based on label
        # What if 'Label' column not present?
        if 'Label' in initial_first_row:
            initial_data_parsed = sorted(initial_data_parsed, key=lambda k: k['Label'] if k['Label'] else "")
        else:
            print("No Label column present, so not sorting this.") #do we need to sort - yes, for diff!

        first_row = row_data_parsed[0]
        header = [k for k in first_row.keys()]
        del header[0]
        # Sort based on label
        # What if 'Label' column not present?
        if 'Label' in first_row:
            row_data_parsed = sorted(row_data_parsed, key=lambda k: k['Label'] if k['Label'] else "")
        else:
            print("No Label column present, so not sorting this.") #do we need to sort - yes, for diff! 

        
        print("Got file_sha",file_sha)

        wb = openpyxl.Workbook()
        sheet = wb.active

        for c in range(len(header)):
            sheet.cell(row=1, column=c+1).value=header[c]
            sheet.cell(row=1, column=c+1).font = Font(size=12,bold=True)
        for r in range(len(row_data_parsed)):
            row = [v for v in row_data_parsed[r].values()]
            del row[0] # Tabulator-added ID column
            for c in range(len(header)):
                sheet.cell(row=r+2, column=c+1).value=row[c]
                # Set row background colours according to 'Curation status'
                # These should be kept in sync with those used in edit screen
                # TODO add to config
                # What if "Curation status" not present?
                if 'Curation status' in first_row:
                    if row[header.index("Curation status")]=="Discussed":
                        sheet.cell(row=r+2, column=c+1).fill = PatternFill(fgColor="ffe4b5", fill_type = "solid")
                    elif row[header.index("Curation status")]=="Ready": #this is depreciated
                        sheet.cell(row=r+2, column=c+1).fill = PatternFill(fgColor="98fb98", fill_type = "solid")
                    elif row[header.index("Curation status")]=="Proposed":
                        sheet.cell(row=r+2, column=c+1).fill = PatternFill(fgColor="ffffff", fill_type = "solid")
                    elif row[header.index("Curation status")]=="To Be Discussed":
                        sheet.cell(row=r+2, column=c+1).fill = PatternFill(fgColor="eee8aa", fill_type = "solid")
                    elif row[header.index("Curation status")]=="In Discussion":
                        sheet.cell(row=r+2, column=c+1).fill = PatternFill(fgColor="fffacd", fill_type = "solid")                                
                    elif row[header.index("Curation status")]=="Published":
                        sheet.cell(row=r+2, column=c+1).fill = PatternFill(fgColor="7fffd4", fill_type = "solid")
                    elif row[header.index("Curation status")]=="Obsolete":
                        sheet.cell(row=r+2, column=c+1).fill = PatternFill(fgColor="2f4f4f", fill_type = "solid")

            # Generate identifiers:
            if 'ID' in first_row:             
                if not row[header.index("ID")]: #blank
                    if 'Label' and 'Parent' and 'Definition' in first_row: #make sure we have the right sheet
                        if row[header.index("Label")] and row[header.index("Parent")] and row[header.index("Definition")]: #not blank
                            #generate ID here: 
                            nextIdStr = str(searcher.getNextId(repo_key))
                            fill_num = app.config['DIGIT_COUNT']
                            if repo_key == "BCIO":
                                fill_num = fill_num-1
                            else:
                                fill_num = fill_num
                            id = repo_key.upper()+":"+nextIdStr.zfill(fill_num)
                            new_id = id
                            for c in range(len(header)):
                                if c==0:
                                    restart = True
                                    sheet.cell(row=r+2, column=c+1).value=new_id

        # Create version for saving
        spreadsheet_stream = io.BytesIO()
        wb.save(spreadsheet_stream)

        #base64_bytes = base64.b64encode(sample_string_bytes)
        base64_bytes = base64.b64encode(spreadsheet_stream.getvalue())
        base64_string = base64_bytes.decode("ascii")

        # Create a new branch to commit the change to (in case of simultaneous updates)
        response = github.get(f"repos/{repo_detail}/git/ref/heads/master")
        if not response or "object" not in response or "sha" not in response["object"]:
            raise Exception(f"Unable to get SHA for HEAD of master in {repo_detail}")
        sha = response["object"]["sha"]
        branch = f"{g.user.github_login}_{datetime.utcnow().strftime('%Y-%m-%d_%H%M%S')}"
        print("About to try to create branch in ",f"repos/{repo_detail}/git/refs")
        response = github.post(
            f"repos/{repo_detail}/git/refs", data={"ref": f"refs/heads/{branch}", "sha": sha},
            )
        if not response:
            raise Exception(f"Unable to create new branch {branch} in {repo_detail}")

        print("About to get latest version of the spreadsheet file",f"repos/{repo_detail}/contents/{folder}/{spreadsheet}")
        # Get the sha for the file
        (new_file_sha, new_rows, new_header) = get_spreadsheet(repo_detail,folder, spreadsheet)

        # Commit changes to branch (replace code with sheet)
        data = {
            "message": commit_msg,
            "content": base64_string,
            "branch": branch,
        }
        data["sha"] = new_file_sha
        print("About to commit file to branch",f"repos/{repo_detail}/contents/{folder}/{spreadsheet}")
        response = github.put(f"repos/{repo_detail}/contents/{folder}/{spreadsheet}", data=data)
        if not response:
            raise Exception(
                f"Unable to commit addition of {spreadsheet} to branch {branch} in {repo_detail}"
            )

        # Create a PR for the change
        print("About to create PR from branch",)
        response = github.post(
            f"repos/{repo_detail}/pulls",
            data={
                "title": commit_msg,
                "head": branch,
                "base": "master",
                "body": commit_msg_extra
            },
        )
        if not response:
            raise Exception(f"Unable to create PR for branch {branch} in {repo_detail}")
        pr_info = response['html_url']

        # Do not merge automatically if this file was stale as that will overwrite the other changes
        
        if new_file_sha != file_sha and not overwrite:
            print("PR created and must be merged manually as repo file had changed")

            # Get the changes between the new file and this one:
            merge_diff, merged_table = getDiff(row_data_parsed, new_rows, new_header, initial_data_parsed) # getDiff(saving version, latest server version, header for both)
            # update rows for comparison:
            (file_sha3,rows3,header3) = get_spreadsheet(repo_detail,folder,spreadsheet)
            #todo: delete transient branch here? Github delete code is a test for now. 
            # Delete the branch again
            print ("About to delete branch",f"repos/{repo_detail}/git/refs/heads/{branch}")
            response = github.delete(
                f"repos/{repo_detail}/git/refs/heads/{branch}")
            if not response:
                raise Exception(f"Unable to delete branch {branch} in {repo_detail}")
            return(
                json.dumps({'Error': 'Your change was submitted to the repository but could not be automatically merged due to a conflict. You can view the change <a href="'\
                    + pr_info + '" target = "_blank" >here </a>. ', "file_sha_1": file_sha, "file_sha_2": new_file_sha, "pr_branch":branch, "merge_diff":merge_diff, "merged_table":json.dumps(merged_table),\
                        "rows3": rows3, "header3": header3}), 300 #400 for missing REPO
                )
        else:
            # Merge the created PR
            print("About to merge created PR")
            response = github.post(
                f"repos/{repo_detail}/merges",
                data={
                    "head": branch,
                    "base": "master",
                    "commit_message": commit_msg
                },
            )
            if not response:
                raise Exception(f"Unable to merge PR from branch {branch} in {repo_detail}")

            # Delete the branch again
            print ("About to delete branch",f"repos/{repo_detail}/git/refs/heads/{branch}")
            response = github.delete(
                f"repos/{repo_detail}/git/refs/heads/{branch}")
            if not response:
                raise Exception(f"Unable to delete branch {branch} in {repo_detail}")

        print ("Save succeeded.")
        # Update the search index for this file ASYNCHRONOUSLY (don't wait)
        thread = threading.Thread(target=searcher.updateIndex,
                                  args=(repo_key, folder, spreadsheet, header, row_data_parsed))
        thread.daemon = True  # Daemonize thread
        thread.start()  # Start the execution

        # Get the sha AGAIN for the file
        response = github.get(f"repos/{repo_detail}/contents/{folder}/{spreadsheet}")
        if not response or "sha" not in response:
            raise Exception(
                f"Unable to get the newly updated SHA value for {spreadsheet} in {repo_detail}/{folder}"
                )
        new_file_sha = response['sha']
        if restart: #todo: does this need to be anywhere else also?
            return ( json.dumps({"message":"Success",
                                "file_sha": new_file_sha}), 360 )
        else:
            return ( json.dumps({"message":"Success",
                                "file_sha": new_file_sha}), 200 )

    except Exception as err:
        print(err)
        traceback.print_exc()
        return (
            json.dumps({"message": "Failed",
                        "Error":format(err)}),
            400,
        )



@app.route('/keepalive', methods=['POST'])
@verify_logged_in
def keep_alive():
    #print("Keep alive requested from edit screen")
    return ( json.dumps({"message":"Success"}), 200 )


#todo: use this function to compare initial spreadsheet to server version - check for updates?
@app.route("/checkForUpdates", methods=["POST"])
def checkForUpdates():
    if request.method == "POST":
        repo_key = request.form.get("repo_key") 
        folder = request.form.get("folder")
        spreadsheet = request.form.get("spreadsheet")
        # initialData = request.form.get("initialData") 
        old_sha = request.form.get("file_sha")     
        repositories = app.config['REPOSITORIES']
        repo_detail = repositories[repo_key]
        spreadsheet_file = github.get(
            f'repos/{repo_detail}/contents/{folder}/{spreadsheet}'
        )
        file_sha = spreadsheet_file['sha']
        if old_sha == file_sha:
            return ( json.dumps({"message":"Success"}), 200 )
        else:
            return ( json.dumps({"message":"Fail"}), 200 )

@app.route('/openVisualiseAcrossSheets', methods=['POST'])
@verify_logged_in
def openVisualiseAcrossSheets():
    #build data we need for dotStr query (new one!)
    if request.method == "POST":
        idString = request.form.get("idList")
        repo = request.form.get("repo") 
        idList = idString.split()
        ontodb.parseRelease(repo) 
        #todo: do we need to support more than one repo at a time here?
        dotStr = ontodb.getDotForIDs(repo,idList).to_string()
        return render_template("visualise.html", sheet="selection", repo=repo, dotStr=dotStr)

    return ("Only POST allowed.")


#api:

@app.route('/api/openVisualiseAcrossSheets', methods=['POST'])
# @verify_logged_in # not enabled for /api/
def apiOpenVisualiseAcrossSheets():
    #build data we need for dotStr query (new one!)
    if request.method == "POST":
        idString = request.form.get("idList")
        repo = request.form.get("repo")
        idList = idString.split()           
        ontodb.parseRelease(repo)
        if len(idList) == 0:
            print("got zero length idList")
            allIds = ontodb.getReleaseIDs(repo)
            idList = []
            for ID in allIds: 
                if ID is not None and ID != "":
                    idList.append(ID.strip())
        # print(idList)
        #test from retReleaseIDs:
        # idList = ['DOID:150', 'ADDICTO:0000975', 'GSSO:003558', 'CHEBI:23367', 'CHEBI:60911', 'CHEBI:59999', 'ADDICTO:0000978', 'ADDICTO:0001066', 'CHEBI:15734', 'ADDICTO:0000375', 'CHEBI:33308', 'OBI:0000071', 'ADDICTO:0000124', 'ADDICTO:0000875', 'ADDICTO:0000281', 'ADDICTO:0000495', 'ADDICTO:0001114', 'ADDICTO:0000500', 'ADDICTO:0000296', 'CHEBI:69478', 'CHEBI:33671', 'ADDICTO:0000781', 'ADDICTO:0000943', 'CHEBI:33635', 'ADDICTO:0000353', 'ADDICTO:0000762', 'ADDICTO:0001127', 'ADDICTO:0000873', 'ADDICTO:0000471', 'ADDICTO:0000753', 'GSSO:003501', 'CHEBI:66964', 'ADDICTO:0000678', 'CHEBI:35701', 'ADDICTO:0000733', 'ADDICTO:0000641', 'CHEBI:33832', 'ADDICTO:0000260', 'ADDICTO:0000720', 'ADDICTO:0000886', 'ADDICTO:0000313', 'MF:0000033', 'BFO:0000020', 'ADDICTO:0000285', 'IAO:0000033', 'ADDICTO:0000272', 'ADDICTO:0000896', 'ADDICTO:0000744', 'ADDICTO:0000106', 'CHEBI:3219', 'ADDICTO:0000399', 'ADDICTO:0000263', 'ADDICTO:0000416', 'ADDICTO:0001040', 'ADDICTO:0000880', 'ADDICTO:0000123', 'OGMS:0000102', 'ADDICTO:0000223', 'UBERON:0000105', 'ADDICTO:0000424', 'ADDICTO:0000387', 'ADDICTO:0000941', 'UBERON:0000463', 'ADDICTO:0000303', 'ADDICTO:0000284', 'ADDICTO:0001102', 'ADDICTO:0000930', 'CHEBI:35474', 'ADDICTO:0000120', 'ADDICTO:0000811', 'ADDICTO:0000747', 'ADDICTO:0001070', 'ADDICTO:0000449', 'ADDICTO:0000195', 'ADDICTO:0000205', 'ADDICTO:0001010', 'ADDICTO:0000895', 'ADDICTO:0000672', 'ADDICTO:0000640', 'ADDICTO:0001042', 'ADDICTO:0000764', 'ADDICTO:0000225', 'ADDICTO:0001017', 'ADDICTO:0000716', 'ADDICTO:0001071', 'ADDICTO:0001087', 'GSSO:000130', 'ADDICTO:0000110', 'ADDICTO:0001011', 'ADDICTO:0000432', 'ADDICTO:0000388', 'GSSO:000370', 'BFO:0000040', 'ADDICTO:0000183', 'ADDICTO:0000659', 'ADDICTO:0000377', 'ADDICTO:0001138', 'ADDICTO:0000130', 'ADDICTO:0001074', 'ADDICTO:0000867', 'CHEBI:38166', 'CHEBI:30879', 'ADDICTO:0000405', 'ADDICTO:0000536', 'ADDICTO:0000203', 'CHEBI:35617', 'ADDICTO:0001034', 'ADDICTO:0001029', 'ADDICTO:0000254', 'ADDICTO:0000987', 'ADDICTO:0000288', 'ADDICTO:0000131', 'CHEBI:25806', 'ADDICTO:0000750', 'ADDICTO:0000520', 'ADDICTO:0000980', 'GSSO:000395', 'CHEBI:17478', 'ENVO:01000838', 'ADDICTO:0000855', 'ADDICTO:0000198', 'ADDICTO:0000145', 'ADDICTO:0000118', 'ADDICTO:0000114', 'DOID:399', 'ADDICTO:0000242', 'ADDICTO:0001027', 'ADDICTO:0000230', 'CHEBI:35482', 'ADDICTO:0000315', 'ADDICTO:0000530', 'ADDICTO:0000931', 'OMRSE:00000062', 'ADDICTO:0001088', 'ADDICTO:0000713', 'ADDICTO:0000788', 'ADDICTO:0000837', 'ADDICTO:0001063', 'ADDICTO:0001080', 'ADDICTO:0000437', 'ADDICTO:0000385', 'ENVO:03000043', 'UBERON:0005162', 'MFOEM:000006', 'PATO:0001018', 'ADDICTO:0000323', 'DOID:4', 'ADDICTO:0000153', 'ADDICTO:0000174', 'UBERON:0000062', 'ADDICTO:0000194', 'ADDICTO:0000739', 'ADDICTO:0000343', 'ADDICTO:0000807', 'CHEBI:35476', 'ADDICTO:0000940', 'CHEBI:35469', 'ADDICTO:0000219', 'ADDICTO:0000258', 'ADDICTO:0000456', 'CHEBI:35803', 'CHEBI:73413', 'GSSO:009381', 'ADDICTO:0000683', 'ADDICTO:0001051', 'ADDICTO:0000529', 'GSSO:002957', 'ADDICTO:0000810', 'CHEBI:42797', 'ADDICTO:0000816', 'ADDICTO:0000326', 'ADDICTO:0000396', 'UBERON:0002308', 'UBERON:0001062', 'ADDICTO:0000783', 'ADDICTO:0000628', 'ADDICTO:0000301', 'ADDICTO:0000528', 'ADDICTO:0000139', 'APOLLO_SV:00000306', 'ADDICTO:0001013', 'ADDICTO:0000197', 'ENVO:00010505', 'ADDICTO:0000844', 'ADDICTO:0000645', 'ADDICTO:0001023', 'ADDICTO:0000893', 'CHEBI:50269', 'ADDICTO:0000206', 'ADDICTO:0000993', 'ADDICTO:0000732', 'ENVO:01001654', 'ADDICTO:0000216', 'ADDICTO:0001101', 'ADDICTO:0000676', 'MF:0000016', 'ADDICTO:0001059', 'ADDICTO:0000156', 'ADDICTO:0000249', 'CHEBI:27732', 'ADDICTO:0000187', 'ADDICTO:0001131', 'ADDICTO:0000213', 'GSSO:004516', 'ADDICTO:0000835', 'UBERON:0000465', 'APOLLO_SV:00000298', 'ADDICTO:0000793', 'ADDICTO:0000636', 'OBI:0000423', 'ADDICTO:0000170', 'ADDICTO:0000630', 'ADDICTO:0000898', 'ADDICTO:0000415', 'ADDICTO:0000360', 'ADDICTO:0000324', 'CHEBI:22712', 'ADDICTO:0000531', 'ADDICTO:0000823', 'ADDICTO:0000974', 'CHEBI:17087', 'CHEBI:78840', 'ADDICTO:0000465', 'ADDICTO:0000755', 'ADDICTO:0000328', 'ADDICTO:0000675', 'CHEBI:39106', 'ADDICTO:0000250', 'OBCS:0000071', 'ADDICTO:0000266', 'ADDICTO:0001130', 'ADDICTO:0000460', 'ADDICTO:0000818', 'ADDICTO:0000430', 'ADDICTO:0000257', 'ADDICTO:0000491', 'ADDICTO:0000105', 'ADDICTO:0000188', 'DOID:1561', 'ADDICTO:0000322', 'ADDICTO:0000171', 'ADDICTO:0000354', 'ADDICTO:0000797', 'ENVO:00010483', 'ADDICTO:0000922', 'ADDICTO:0000150', 'ADDICTO:0000237', 'ADDICTO:0000127', 'MFOEM:000001', 'GSSO:009846', 'CHEBI:33597', 'CHEBI:50995', 'BFO:0000141', 'ADDICTO:0001058', 'ADDICTO:0000719', 'ADDICTO:0000759', 'ADDICTO:0000824', 'OBCS:0000220', 'ADDICTO:0000300', 'UBERON:0001873', 'ADDICTO:0000709', 'ADDICTO:0000649', 'CHEBI:22720', 'ADDICTO:0000888', 'ADDICTO:0000282', 'MF:0000031', 'ADDICTO:0001119', 'ADDICTO:0000995', 'ADDICTO:0000684', 'ADDICTO:0000532', 'BFO:0000017', 'ADDICTO:0000176', 'DOID:0060038', 'UBERON:0004535', 'ADDICTO:0000802', 'ADDICTO:0001075', 'CHEBI:27808', 'ADDICTO:0000293', 'ADDICTO:0000938', 'ADDICTO:0000642', 'BCIO:037000', 'ADDICTO:0000348', 'ADDICTO:0000220', 'ADDICTO:0001122', 'ADDICTO:0000479', 'UBERON:0000071', 'ADDICTO:0000839', 'ADDICTO:0000661', 'ADDICTO:0000694', 'CHEBI:33598', 'ADDICTO:0001093', 'ADDICTO:0000971', 'ADDICTO:0000344', 'CHEBI:24532', 'BCIO:042000', 'ADDICTO:0000710', 'CHEBI:23888', 'ADDICTO:0000381', 'ADDICTO:0000132', 'ADDICTO:0000740', 'BFO:0000015', 'BFO:0000023', 'ADDICTO:0000425', 'ADDICTO:0000745', 'ADDICTO:0000634', 'ADDICTO:0000398', 'MFOEM:000195', 'ADDICTO:0000988', 'BCIO:041000', 'ADDICTO:0001069', 'ADDICTO:0000865', 'ADDICTO:0000346', 'ADDICTO:0000992', 'ADDICTO:0000657', 'CHEBI:33232', 'ADDICTO:0000370', 'ADDICTO:0000787', 'IAO:0000178', 'ADDICTO:0000796', 'ADDICTO:0000681', 'ADDICTO:0000341', 'CHEBI:33659', 'ADDICTO:0000447', 'ADDICTO:0000999', 'PO:0025131', 'ADDICTO:0000965', 'ADDICTO:0000717', 'ADDICTO:0001094', 'ADDICTO:0000803', 'PATO:0000033', 'BCIO:036000', 'ADDICTO:0000794', 'ENVO:01000840', 'ADDICTO:0000976', 'ADDICTO:0000466', 'ADDICTO:0000700', 'ADDICTO:0000304', 'ADDICTO:0000889', 'ADDICTO:0000209', 'ADDICTO:0001018', 'ADDICTO:0000160', 'ADDICTO:0000984', 'ADDICTO:0000279', 'ADDICTO:0000478', 'ADDICTO:0000117', 'ADDICTO:0000511', 'ADDICTO:0000401', 'BCIO:003000', 'ADDICTO:0000534', 'ADDICTO:0001038', 'CHEBI:38323', 'CHEBI:27171', 'CHEBI:33636', 'ADDICTO:0000179', 'GSSO:000232', 'SEPIO:0000004', 'ADDICTO:0000842', 'CHEBI:67196', 'ADDICTO:0000773', 'CHEBI:51143', 'ADDICTO:0000443', 'ADDICTO:0000475', 'ADDICTO:0000261', 'ADDICTO:0000391', 'ADDICTO:0001113', 'ADDICTO:0000846', 'ADDICTO:0000164', 'GSSO:000379', 'ADDICTO:0000302', 'ADDICTO:0000366', 'CHEBI:32952', 'ADDICTO:0000368', 'ADDICTO:0000705', 'BFO:0000029', 'CHEBI:16842', 'BFO:0000027', 'CHEBI:33848', 'ADDICTO:0000779', 'CHEBI:16236', 'OMRSE:00000061', 'ADDICTO:0000956', 'ADDICTO:0000420', 'ADDICTO:0000813', 'ADDICTO:0000654', 'ADDICTO:0001020', 'CHEBI:76224', 'ADDICTO:0000703', 'ADDICTO:0000191', 'ADDICTO:0000312', 'CHEBI:33579', 'ADDICTO:0001091', 'ADDICTO:0000854', 'ADDICTO:0000687', 'ADDICTO:0001128', 'CHEBI:27958', 'ADDICTO:0000365', 'ADDICTO:0001060', 'ADDICTO:0000140', 'ADDICTO:0000444', 'ADDICTO:0000234', 'ADDICTO:0000143', 'ADDICTO:0000851', 'ADDICTO:0000955', 'ADDICTO:0000925', 'PATO:0000001', 'ADDICTO:0000459', 'ADDICTO:0000994', 'ADDICTO:0000736', 'CHEBI:33575', 'ADDICTO:0000688', 'ADDICTO:0000670', 'ADDICTO:0000202', 'ADDICTO:0000116', 'ADDICTO:0000843', 'ADDICTO:0000997', 'ADDICTO:0000690', 'ADDICTO:0000652', 'ADDICTO:0000910', 'IAO:0000030', 'CHEBI:48878', 'ADDICTO:0000790', 'ADDICTO:0000292', 'ADDICTO:0000981', 'ADDICTO:0000409', 'ADDICTO:0000708', 'ADDICTO:0000638', 'ADDICTO:0000513', 'CHEBI:34967', 'ADDICTO:0000112', 'CHEBI:41607', 'ADDICTO:0000363', 'ADDICTO:0000653', 'ADDICTO:0000725', 'ADDICTO:0000172', 'ADDICTO:0000107', 'ADDICTO:0001065', 'DOID:2030', 'ADDICTO:0000696', 'ADDICTO:0000872', 'ADDICTO:0000715', 'ADDICTO:0000207', 'ADDICTO:0000831', 'ADDICTO:0000362', 'OBI:0000047', 'ADDICTO:0000862', 'ADDICTO:0001115', 'ADDICTO:0000310', 'CHEBI:25418', 'ADDICTO:0000294', 'ADDICTO:0000464', 'ADDICTO:0000903', 'ADDICTO:0000241', 'ADDICTO:0000325', 'UBERON:0000125', 'ADDICTO:0001028', 'ADDICTO:0000632', 'ADDICTO:0000920', 'OBCS:0000218', 'ADDICTO:0000327', 'UBERON:0000000', 'CHEBI:84500', 'ADDICTO:0000891', 'ADDICTO:0000364', 'CHEBI:36586', 'ADDICTO:0000729', 'ADDICTO:0000278', 'UBERON:0000477', 'ADDICTO:0000305', 'ADDICTO:0001125', 'ADDICTO:0000874', 'ADDICTO:0000269', 'ADDICTO:0000658', 'ADDICTO:0000469', 'ADDICTO:0001124', 'ADDICTO:0000932', 'ADDICTO:0000686', 'ADDICTO:0000808', 'ADDICTO:0001000', 'ADDICTO:0000780', 'ADDICTO:0000866', 'ADDICTO:0000801', 'ADDICTO:0000245', 'CHEBI:52210', 'ADDICTO:0000392', 'CHEBI:37622', 'ADDICTO:0000196', 'ADDICTO:0000321', 'CHEBI:3216', 'ENVO:01001222', 'GSSO:000498', 'ADDICTO:0000669', 'ADDICTO:0001021', 'UBERON:0000061', 'ADDICTO:0000990', 'ADDICTO:0000644', 'ADDICTO:0000926', 'ADDICTO:0000977', 'ADDICTO:0000714', 'ADDICTO:0000936', 'ADDICTO:0000923', 'ADDICTO:0000470', 'ADDICTO:0000699', 'ADDICTO:0000728', 'ADDICTO:0000109', 'ADDICTO:0000778', 'CHEBI:6809', 'ADDICTO:0000772', 'ADDICTO:0000635', 'CHEBI:50906', 'GSSO:004615', 'OBI:0000011', 'CHEBI:33637', 'ADDICTO:0000244', 'ADDICTO:0001117', 'ADDICTO:0000247', 'ADDICTO:0000154', 'CHEBI:83818', 'ADDICTO:0000215', 'ADDICTO:0000421', 'ADDICTO:0000624', 'CHEBI:16482', 'ADDICTO:0000218', 'ADDICTO:0000290', 'ADDICTO:0001015', 'ADDICTO:0001100', 'ADDICTO:0000737', 'CHEBI:33304', 'ADDICTO:0000355', 'CHEBI:37577', 'ADDICTO:0000236', 'ADDICTO:0000283', 'ADDICTO:0000730', 'ADDICTO:0001118', 'BCIO:038000', 'ADDICTO:0000968', 'CHEBI:50996', 'BFO:0000004', 'ADDICTO:0001048', 'ENVO:01000839', 'ADDICTO:0001077', 'ADDICTO:0000775', 'ADDICTO:0000689', 'ADDICTO:0000841', 'ADDICTO:0000402', 'ADDICTO:0000677', 'ADDICTO:0001081', 'DOID:3324', 'ADDICTO:0000754', 'ADDICTO:0001045', 'CHEBI:17245', 'ADDICTO:0001064', 'ADDICTO:0000457', 'ADDICTO:0000718', 'ADDICTO:0001089', 'ADDICTO:0000246', 'CHEBI:33822', 'ADDICTO:0000514', 'ADDICTO:0000157', 'ADDICTO:0000998', 'CHEBI:7459', 'ADDICTO:0000451', 'ADDICTO:0000351', 'DOID:0050117', 'ADDICTO:0000827', 'ADDICTO:0001001', 'ADDICTO:0000890', 'ADDICTO:0000135', 'ADDICTO:0000760', 'ADDICTO:0000185', 'ADDICTO:0000706', 'ADDICTO:0000735', 'ADDICTO:0000782', 'CHEBI:17153', 'ADDICTO:0000199', 'ADDICTO:0000101', 'ADDICTO:0000982', 'CHEBI:35473', 'ADDICTO:0000897', 'ADDICTO:0000200', 'CHEBI:24471', 'ADDICTO:0000795', 'ADDICTO:0000655', 'BCIO:034000', 'ADDICTO:0000267', 'ADDICTO:0000734', 'ADDICTO:0000944', 'ADDICTO:0000738', 'ADDICTO:0001044', 'ADDICTO:0000685', 'ADDICTO:0000809', 'ADDICTO:0000119', 'ADDICTO:0001085', 'ERO:0001108', 'ADDICTO:0000253', 'ADDICTO:0001008', 'ADDICTO:0000308', 'ADDICTO:0000488', 'ADDICTO:0001112', 'ADDICTO:0000538', 'ADDICTO:0001005', 'ADDICTO:0000523', 'BFO:0000016', 'ADDICTO:0000276', 'ADDICTO:0001049', 'ADDICTO:0000757', 'ADDICTO:0001078', 'CHEBI:35352', 'ADDICTO:0000255', 'ADDICTO:0000271', 'CHEBI:51177', 'ADDICTO:0000129', 'ADDICTO:0000412', 'ADDICTO:0001092', 'ADDICTO:0000359', 'ADDICTO:0000800', 'ADDICTO:0000524', 'ADDICTO:0000884', 'CHEBI:2972', 'ENVO:00002221', 'BCIO:043000', 'GSSO:000376', 'RO:0002577', 'ADDICTO:0000204', 'ADDICTO:0000919', 'ADDICTO:0000458', 'ADDICTO:0000861', 'CHEBI:5686', 'ADDICTO:0001136', 'ADDICTO:0000881', 'ADDICTO:0000525', 'CHEBI:25693', 'ADDICTO:0000320', 'ADDICTO:0000752', 'ADDICTO:0000349', 'CHEBI:33285', 'ADDICTO:0000192', 'IAO:0000027', 'ADDICTO:0000180', 'ADDICTO:0000201', 'ADDICTO:0001016', 'CHEBI:7852', 'CHEBI:49575', 'GSSO:005379', 'ADDICTO:0001098', 'CHMO:0001000', 'CHEBI:33833', 'UBERON:0035943', 'ADDICTO:0000445', 'ADDICTO:0000516', 'ADDICTO:0000799', 'ADDICTO:0000299', 'ADDICTO:0000665', 'ADDICTO:0001111', 'ADDICTO:0000776', 'ADDICTO:0000674', 'ADDICTO:0001002', 'IAO:0000088', 'ADDICTO:0000248', 'CHEBI:67194', 'ADDICTO:0000633', 'GSSO:001596', 'ADDICTO:0000834', 'ADDICTO:0000829', 'CHEBI:35990', 'ADDICTO:0001083', 'ADDICTO:0000985', 'UBERON:0010000', 'OBCS:0000150', 'DOID:526', 'ADDICTO:0001073', 'ADDICTO:0000857', 'CHEBI:73416', 'CHEBI:36963', 'ADDICTO:0000887', 'ADDICTO:0000316', 'ADDICTO:0000860', 'ADDICTO:0000222', 'ADDICTO:0000832', 'ADDICTO:0000345', 'ADDICTO:0000406', 'GSSO:002821', 'ADDICTO:0001032', 'ADDICTO:0000492', 'ADDICTO:0000347', 'ADDICTO:0000390', 'CHEBI:2679', 'ADDICTO:0000820', 'ADDICTO:0000527', 'BFO:0000003', 'ADDICTO:0000251', 'GSSO:005835', 'MF:0000020', 'GSSO:001802', 'ADDICTO:0000168', 'ADDICTO:0000849', 'DOID:0050668', 'ADDICTO:0000784', 'ADDICTO:0001072', 'ADDICTO:0001061', 'ADDICTO:0000702', 'ADDICTO:0000311', 'CHEBI:50860', 'ADDICTO:0000450', 'ADDICTO:0001110', 'ADDICTO:0001012', 'ADDICTO:0000136', 'ADDICTO:0001129', 'ADDICTO:0000771', 'ADDICTO:0000232', 'UBERON:0006314', 'ADDICTO:0000186', 'ENVO:00000073', 'ADDICTO:0000918', 'ADDICTO:0000693', 'ADDICTO:0000991', 'ADDICTO:0000357', 'GSSO:002024', 'ADDICTO:0000969', 'ADDICTO:0000411', 'ADDICTO:0000350', 'MF:0000030', 'ENVO:01000786', 'ADDICTO:0000155', 'ADDICTO:0000193', 'ADDICTO:0000427', 'GSSO:001595', 'ADDICTO:0001062', 'ADDICTO:0000422', 'ADDICTO:0001103', 'ADDICTO:0001090', 'ADDICTO:0000468', 'CHEBI:73417', 'ADDICTO:0000142', 'ADDICTO:0000126', 'ADDICTO:0000490', 'CHEBI:1391', 'ADDICTO:0000983', 'ADDICTO:0001019', 'GSSO:002961', 'CHEBI:38104', 'CHEBI:7465', 'CHEBI:35338', 'ADDICTO:0000461', 'ADDICTO:0000314', 'CHEBI:37249', 'UBERON:0011215', 'ADDICTO:0000650', 'ADDICTO:0000939', 'ADDICTO:0000701', 'CHEBI:33675', 'ADDICTO:0000850', 'ADDICTO:0000356', 'GSSO:000229', 'ADDICTO:0000374', 'BFO:0000031', 'CHEBI:38164', 'GSSO:007328', 'ADDICTO:0000133', 'ADDICTO:0000666', 'ADDICTO:0000414', 'ADDICTO:0000438', 'ADDICTO:0000533', 'ADDICTO:0000828', 'ADDICTO:0000517', 'ADDICTO:0000178', 'ADDICTO:0000435', 'ADDICTO:0000413', 'ADDICTO:0000397', 'ADDICTO:0001009', 'ADDICTO:0000372', 'CHEBI:24651', 'ADDICTO:0001014', 'ADDICTO:0000108', 'ADDICTO:0001036', 'ADDICTO:0000515', 'ADDICTO:0000870', 'BFO:0000024', 'ADDICTO:0001003', 'ADDICTO:0000656', 'ADDICTO:0000927', 'CHEBI:36357', 'ADDICTO:0001047', 'ADDICTO:0000805', 'CHEBI:24432', 'ADDICTO:0000115', 'ADDICTO:0000395', 'ADDICTO:0001106', 'ADDICTO:0000400', 'ADDICTO:0000964', 'ADDICTO:0000821', 'CHEBI:127342', 'ADDICTO:0000966', 'ADDICTO:0001137', 'ADDICTO:0000830', 'CHEBI:33595', 'ADDICTO:0000455', 'ADDICTO:0000394', 'ADDICTO:0001139', 'ADDICTO:0000660', 'ADDICTO:0000221', 'ADDICTO:0000731', 'CHEBI:26979', 'SEPIO:0000125', 'CHEBI:50903', 'ADDICTO:0000958', 'ENVO:00000070', 'ADDICTO:0000777', 'ADDICTO:0000104', 'ADDICTO:0000847', 'ADDICTO:0001116', 'ADDICTO:0000815', 'ADDICTO:0001025', 'CHMO:0000999', 'CHMO:0001004', 'ADDICTO:0000774', 'ADDICTO:0001121', 'ADDICTO:0001082', 'ADDICTO:0000477', 'ADDICTO:0000378', 'OGMS:0000045', 'ADDICTO:0000417', 'ADDICTO:0000751', 'ADDICTO:0000756', 'ADDICTO:0000137', 'ADDICTO:0001035', 'CHEBI:18723', 'ADDICTO:0001076', 'ADDICTO:0000173', 'ADDICTO:0000442', 'ADDICTO:0000512', 'ADDICTO:0000804', 'ADDICTO:0000240', 'ADDICTO:0000410', 'ADDICTO:0000648', 'DOID:2468', 'ADDICTO:0000373', 'ADDICTO:0000626', 'ADDICTO:0000147', 'ADDICTO:0000231', 'ADDICTO:0000637', 'ADDICTO:0000639', 'ADDICTO:0000743', 'ADDICTO:0000791', 'BFO:0000144', 'ADDICTO:0000631', 'CHEBI:83403', 'ADDICTO:0001123', 'ADDICTO:0000712', 'ADDICTO:0000871', 'DOID:12995', 'ADDICTO:0000227', 'PATO:0002182', 'ADDICTO:0000883', 'ADDICTO:0000431', 'ADDICTO:0001039', 'OMRSE:00000114', 'CHEBI:36683', 'ADDICTO:0000767', 'ADDICTO:0001086', 'ADDICTO:0000825', 'ADDICTO:0000963', 'MFOEM:000005', 'ADDICTO:0000141', 'ADDICTO:0000786', 'ADDICTO:0000167', 'ADDICTO:0000151', 'ADDICTO:0000996', 'ADDICTO:0000175', 'CHEBI:72695', 'ADDICTO:0000798', 'ADDICTO:0000765', 'CHEBI:33663', 'ADDICTO:0001007', 'ADDICTO:0000697', 'ADDICTO:0000858', 'ADDICTO:0000428', 'OMRSE:00000102', 'ADDICTO:0001120', 'ADDICTO:0000214', 'ADDICTO:0000352', 'ADDICTO:0000768', 'ADDICTO:0000217', 'ADDICTO:0000848', 'ADDICTO:0000121', 'ADDICTO:0000907', 'ADDICTO:0000439', 'ADDICTO:0000748', 'CHEBI:47958', 'CHEBI:52217', 'ADDICTO:0000419', 'ADDICTO:0000806', 'ADDICTO:0001057', 'ADDICTO:0000667', 'ENVO:02500000', 'UBERON:0000467', 'ADDICTO:0000190', 'GSSO:002962', 'BFO:0000008', 'ADDICTO:0000434', 'OGMS:0000087', 'ADDICTO:0000822', 'ADDICTO:0001043', 'ADDICTO:0000436', 'ADDICTO:0000664', 'ADDICTO:0000229', 'ADDICTO:0000383', 'CHEBI:32988', 'GSSO:001590', 'GSSO:000529', 'ADDICTO:0000389', 'ADDICTO:0000211', 'ADDICTO:0000149', 'ADDICTO:0000651', 'ADDICTO:0000704', 'CHEBI:50047', 'ADDICTO:0000342', 'ADDICTO:0000673', 'CHEBI:24431', 'CHEBI:35477', 'ADDICTO:0000408', 'ADDICTO:0000268', 'ADDICTO:0000252', 'ADDICTO:0000646', 'ADDICTO:0000679', 'ADDICTO:0000863', 'CHEBI:64708', 'ADDICTO:0000358', 'IAO:0000007', 'ADDICTO:0000960', 'OMRSE:00000150', 'ADDICTO:0000487', 'ADDICTO:0000146', 'ADDICTO:0000440', 'DOID:162', 'ADDICTO:0000361', 'ADDICTO:0000724', 'ADDICTO:0000407', 'ADDICTO:0000518', 'CHEBI:67072', 'ADDICTO:0001133', 'UBERON:0001007', 'ADDICTO:0000233', 'ADDICTO:0000138', 'OBCS:0000035', 'IAO:0000109', 'CHEBI:32692', 'ADDICTO:0000668', 'ENVO:01001813', 'ADDICTO:0001105', 'ADDICTO:0000273', 'CHEBI:22315', 'OBI:0000094', 'ADDICTO:0000535', 'ADDICTO:0000973', 'ADDICTO:0001055', 'ADDICTO:0000102', 'ADDICTO:0000177', 'ADDICTO:0000289', 'ADDICTO:0001004', 'ADDICTO:0000929', 'ADDICTO:0000386', 'ADDICTO:0001079', 'ADDICTO:0000184', 'ADDICTO:0000663', 'BFO:0000030', 'CHEBI:33836', 'DOID:1510', 'ADDICTO:0000691', 'ADDICTO:0000403', 'MF:0000032', 'ADDICTO:0000882', 'ADDICTO:0000239', 'ADDICTO:0001056', 'ADDICTO:0001031', 'ADDICTO:0000695', 'ADDICTO:0000128', 'ADDICTO:0000629', 'ADDICTO:0000122', 'ADDICTO:0000962', 'ADDICTO:0000404', 'ADDICTO:0000647', 'CHEBI:119915', 'ADDICTO:0000474', 'ADDICTO:0001109', 'ADDICTO:0000298', 'CHEBI:35488', 'CHEBI:25367', 'ADDICTO:0000111', 'ADDICTO:0000393', 'ADDICTO:0000726', 'ADDICTO:0000671', 'CHMO:0001001', 'ADDICTO:0000967', 'ADDICTO:0000662', 'ADDICTO:0000826', 'ADDICTO:0000721', 'OGMS:0000031', 'DOID:0060903', 'ADDICTO:0001052', 'ADDICTO:0000986', 'ADDICTO:0000295', 'IAO:0000104', 'ADDICTO:0000319', 'CHEBI:6807', 'ADDICTO:0000643', 'DOID:10937', 'ADDICTO:0000433', 'CHEBI:35471', 'ADDICTO:0000682', 'ADDICTO:0000935', 'BCIO:040000', 'ADDICTO:0001053', 'CHEBI:35293', 'ADDICTO:0000904', 'ADDICTO:0001107', 'ADDICTO:0001022', 'ADDICTO:0000727', 'ADDICTO:0000467', 'ADDICTO:0000228', 'OMRSE:00000106', 'ADDICTO:0000814', 'ADDICTO:0000317', 'OBCS:0000233', 'ADDICTO:0000970', 'ADDICTO:0000817', 'ADDICTO:0000429', 'ADDICTO:0000812', 'OBI:0000984', 'ADDICTO:0000166', 'CHEBI:59331', 'ADDICTO:0000627', 'DOID:1470', 'ADDICTO:0000707', 'ADDICTO:0000892', 'ADDICTO:0000212', 'PO:0000003', 'ADDICTO:0001095', 'ADDICTO:0000845', 'ADDICTO:0000692', 'IAO:0000310', 'ADDICTO:0000979', 'CHEBI:36785', 'ADDICTO:0000722', 'ADDICTO:0000885', 'ADDICTO:0001026', 'ADDICTO:0000152', 'ADDICTO:0001006', 'ADDICTO:0000259', 'ADDICTO:0000256', 'ADDICTO:0000625', 'ADDICTO:0000262', 'DOID:10935', 'BFO:0000002', 'ADDICTO:0000972', 'OGMS:0000060', 'ADDICTO:0000210', 'CHEBI:35294', 'CHEBI:5779', 'ADDICTO:0000113', 'BFO:0000035', 'ADDICTO:0001050', 'BFO:0000001', 'CHEBI:33707', 'GSSO:000369', 'ADDICTO:0000224', 'UBERON:0009663', 'ADDICTO:0000819', 'ADDICTO:0000942', 'GSSO:000924', 'ADDICTO:0000243', 'ADDICTO:0000680', 'ADDICTO:0000144', 'ADDICTO:0001135', 'CHEBI:67201', 'DOID:3312', 'ADDICTO:0000769', 'ADDICTO:0000426', 'CHEBI:33608', 'CHEBI:26385', 'OBI:0000245', 'CHEBI:38101', 'PATO:0001241', 'ADDICTO:0000367', 'ADDICTO:0000906', 'ADDICTO:0000723', 'ADDICTO:0000989', 'ADDICTO:0000836', 'ADDICTO:0001132', 'CHEBI:4055', 'ADDICTO:0000770', 'CHEBI:35470', 'ADDICTO:0001099', 'ADDICTO:0000961', 'ADDICTO:0000792', 'ADDICTO:0000959', 'ADDICTO:0000287', 'GSSO:005301', 'ADDICTO:0000446', 'ADDICTO:0000161']
        #test from front end: 
        # idList = ['ADDICTO:0000323', 'ADDICTO:0000322', 'ADDICTO:0000324', 'ADDICTO:0000967', 'ADDICTO:0000325', 'ADDICTO:0000326']
                # print(ID)
        dotStr = ontodb.getDotForIDs(repo,idList).to_string()
        #NOTE: APP_TITLE2 can't be blank - messes up the spacing  
        APP_TITLE2 = "VISUALISATION" #could model this on calling url here? Or something else..
        #test full sheet: 
        return render_template("visualiseapi.html", sheet="selection", repo=repo, dotStr=dotStr, api=True, APP_TITLE2=APP_TITLE2)
        # return render_template("visualise.html", sheet="selection", repo=repo, dotStr=dotStr, api=True, APP_TITLE2=APP_TITLE2)


@app.route('/openVisualise', methods=['POST'])
@verify_logged_in 
def openVisualise():
    curation_status_filters = ["", "External", "Proposed", "To Be Discussed", "In Discussion", "Discussed", "Published", "Obsolete"]
    dotstr_list = []
    if request.method == "POST":
        repo = request.form.get("repo")
        sheet = request.form.get("sheet")
        table = json.loads(request.form.get("table"))
        indices = json.loads(request.form.get("indices"))
        try: 
            filter = json.loads(request.form.get("filter"))
            if len(filter) > 0:
                pass
            else:
                filter = ""
        except Exception as err:
            filter = ""
            print(err)
        if repo not in ontodb.releases:
            ontodb.parseRelease(repo)

        if len(indices) > 0: #visualise selection
            #check if filter is greater than 1:
            if len(filter) > 1 and filter != "": #multi-select:
                for i in range(0,2):
                    ontodb.parseSheetData(repo,table)
                    dotStr = ontodb.getDotForSelection(repo,table,indices, filter).to_string() #filter is a list of strings here
            else:          
                for filter in curation_status_filters:
                    #loop this twice to mitigate ID bug:   
                    for i in range(0,2):
                        ontodb.parseSheetData(repo,table)
                        dotStr = ontodb.getDotForSelection(repo,table,indices, filter).to_string()
                    #append dotStr to dotstr_list   
                    dotstr_list.append(dotStr) #all possible graphs
                #calculate default all values:
                filter = "" #default
                for i in range(0,2):
                    ontodb.parseSheetData(repo,table)
                    dotStr = ontodb.getDotForSelection(repo,table,indices, filter).to_string()
            
        else:
            #check if filter is greater than 1:
            if len(filter) > 1 and filter != "": #multi-select:
                for i in range(0,2):
                    ontodb.parseSheetData(repo,table)
                    dotStr = ontodb.getDotForSheetGraph(repo,table,filter).to_string() #filter is a list of strings here
                    # no dotstr_list here, just one dotStr
                    # dotstr_list.append(dotStr) #all possible graphs
            else:
                for filter in curation_status_filters: #Visualise sheet
                    #loop this twice to mitigate ID bug:   
                    for i in range(0,2):
                        ontodb.parseSheetData(repo,table)
                        dotStr = ontodb.getDotForSheetGraph(repo,table,filter).to_string()
                    #append dotStr to dotstr_list   
                    dotstr_list.append(dotStr) #all possible graphs
                #calculate default all values:
                filter = "" #default
                for i in range(0,2):
                    ontodb.parseSheetData(repo,table)
                    dotStr = ontodb.getDotForSheetGraph(repo,table,filter).to_string()            

        return render_template("visualise.html", sheet=sheet, repo=repo, dotStr=dotStr, dotstr_list=dotstr_list, filter=filter)

    return ("Only POST allowed.")


# todo: below is never reached? 
# @app.route('/visualise/<repo>/<sheet>')
# @verify_logged_in # todo: does this need to be disabled to allow cross origin requests? apparently not!
# def visualise(repo, sheet):
#     # print("reached visualise")
#     return render_template("visualise.html", sheet=sheet, repo=repo)

@app.route('/openPat', methods=['POST'])
@verify_logged_in
def openPat():
    if request.method == "POST":
        repo = request.form.get("repo")
        # print("repo is ", repo)
        sheet = request.form.get("sheet")
        # print("sheet is ", sheet)
        table = json.loads(request.form.get("table"))
        # print("table is: ", table)
        indices = json.loads(request.form.get("indices"))
        # print("indices are: ", indices)

        if repo not in ontodb.releases:
            ontodb.parseRelease(repo)
        if len(indices) > 0: #selection
            # ontodb.parseSheetData(repo,table)
            allIDS = ontodb.getIDsFromSelection(repo,table,indices)
            # print("got allIDS: ", allIDS)
        else: # whole sheet..
            ontodb.parseSheetData(repo,table)
            allIDS = ontodb.getIDsFromSheet(repo, table)
            #todo: do we need to do above twice? 
        
        # print("allIDS: ", allIDS)
        #remove duplicates from allIDS: 
        allIDS = list(dict.fromkeys(allIDS))
        
        allData = ontodb.getMetaData(repo, allIDS)  
        # print("allData: ", allData) 
        # print("dotStr is: ", dotStr)
        return render_template("pat.html", repo=repo, all_ids=allIDS, all_data=allData) #todo: PAT.html

    return ("Only POST allowed.")

@app.route('/openPatAcrossSheets', methods=['POST'])
@verify_logged_in
def openPatAcrossSheets():
    #build data we need for dotStr query (new one!)
    if request.method == "POST":
        idString = request.form.get("idList")
        # print("idString is: ", idString)
        repo = request.form.get("repo") 
        print("repo is ", repo)
        idList = idString.split()
        print("idList: ", idList)
        # for i in idList:
        #     print("i is: ", i)
        # indices = json.loads(request.form.get("indices"))
        # print("indices are: ", indices)
        ontodb.parseRelease(repo)
        #todo: do we need to support more than one repo at a time here?
        allIDS = ontodb.getRelatedIDs(repo,idList)
        # print("allIDS: ", allIDS)
        #remove duplicates from allIDS: 
        allIDS = list(dict.fromkeys(allIDS))

        #todo: all experimental from here: 
        allData = ontodb.getMetaData(repo, allIDS)  
        # print("TEST allData: ", allData)

        # dotStr = ontodb.getDotForIDs(repo,idList).to_string()
        return render_template("pat.html", repo=repo, all_ids=allIDS, all_data=allData) #todo: PAT.html

    return ("Only POST allowed.")

@app.route('/edit_external/<repo_key>/<path:folder_path>')
@verify_logged_in
def edit_external(repo_key, folder_path):
    # print("edit_external reached") 
    repositories = app.config['REPOSITORIES']
    repo_detail = repositories[repo_key]
    folder=folder_path
    spreadsheets = []
    directories = github.get(
        f'repos/{repo_detail}/contents/{folder_path}' 
    )   
    for directory in directories:
        spreadsheets.append(directory['name'])
    #todo: need unique name for each? Or do we append to big array? 
    # for spreadsheet in spreadsheets:
    #     print("spreadsheet: ", spreadsheet)
        
    sheet1, sheet2, sheet3 = spreadsheets
    (file_sha1,rows1,header1) = get_spreadsheet(repo_detail,folder,sheet1)
    # not a spreadsheet but a csv file:
    (file_sha2,rows2,header2) = get_csv(repo_detail,folder,sheet2) 
    (file_sha3,rows3,header3) = get_csv(repo_detail,folder,sheet3)
    return render_template('edit_external.html', 
                            login=g.user.github_login, 
                            repo_name = repo_key,
                            folder_path = folder_path,
                            spreadsheets=spreadsheets, #todo: delete, just for test
                            rows1=json.dumps(rows1),
                            rows2=json.dumps(rows2),
                            rows3=json.dumps(rows3)
                            )

@app.route('/save_new_ontology', methods=['POST'])
@verify_logged_in
def save_new_ontology():
    new_ontology = request.form.get("new_ontology")
    print("Received new Ontology: " + new_ontology)
    response = "test"
    return ( json.dumps({"message":"Success",
                             "response": response}), 200 )

@app.route('/update_ids', methods=['POST'])
@verify_logged_in
def update_ids():
    current_ontology=request.form.get("current_ontology")
    new_IDs=request.form.get("new_IDs")
    print("Received new IDs from : "+ current_ontology)
    print("data is: ", new_IDs)
    response = "test"
    return ( json.dumps({"message":"Success",
                             "response": response}), 200 )

# Internal methods

def get_csv(repo_detail,folder,spreadsheet):

    csv_file = github.get(
        f'repos/{repo_detail}/contents/{folder}/{spreadsheet}'
    )
    file_sha = csv_file['sha']
    csv_content = csv_file['content']
    # print(csv_content)
    decoded_data = str(base64.b64decode(csv_content),'utf-8')
    # print(decoded_data)
    csv_reader = csv.reader(io.StringIO(decoded_data))
    csv_data = list(csv_reader)
    header = csv_data[0:1]
    rows = csv_data[1:]
    
    # print(f'{spreadsheet} header: ', header)
    # print(f'{spreadsheet} rows: ', rows)

    return ( (file_sha, rows, header) )

def get_spreadsheet(repo_detail,folder,spreadsheet):
    spreadsheet_file = github.get(
        f'repos/{repo_detail}/contents/{folder}/{spreadsheet}'
    )
    file_sha = spreadsheet_file['sha']
    base64_bytes = spreadsheet_file['content'].encode('utf-8')
    decoded_data = base64.decodebytes(base64_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(decoded_data))
    sheet = wb.active

    header = [cell.value for cell in sheet[1] if cell.value]
    rows = []
    for row in sheet[2:sheet.max_row]:
        values = {}
        for key, cell in zip(header, row):
            values[key] = cell.value
        if any(values.values()):
            rows.append(values)
    # print(f'rows: ')
    # print(json.dumps(rows))
    return ( (file_sha, rows, header) )


def getDiff(row_data_1, row_data_2, row_header, row_data_3): #(1saving, 2server, header, 3initial)

    # print(f'the type of row_data_3 is: ')
    # print(type(row_data_3))        

    #sort out row_data_1 format to be the same as row_data_2
    new_row_data_1 = []
    for k in row_data_1:
        dictT = {}
        for key, val, item in zip(k, k.values(), k.items()):
            if(key != "id"):
                if(val == ""):
                    val = None
                #add to dictionary:
                dictT[key] = val
        #add to list:
        new_row_data_1.append( dictT ) 

    #sort out row_data_3 format to be the same as row_data_2
    new_row_data_3 = []
    for h in row_data_3:
        dictT3 = {}
        for key, val, item in zip(h, h.values(), h.items()):
            if(key != "id"):
                if(val == ""):
                    val = None
                #add to dictionary:
                dictT3[key] = val
        #add to list:
        new_row_data_3.append( dictT3 ) 

    row_data_combo_1 = [row_header] 
    row_data_combo_2 = [row_header]
    row_data_combo_3 = [row_header]

    row_data_combo_1.extend([list(r.values()) for r in new_row_data_1]) #row_data_1 has extra "id" column for some reason???!!!
    row_data_combo_2.extend([list(s.values()) for s in row_data_2])
    row_data_combo_3.extend([list(t.values()) for t in new_row_data_3])

    #checking:
    # print(f'row_header: ')
    # print(row_header)
    # print(f'row_data_1: ')
    # print(row_data_1)
    # print(f'row_data_2: ')
    # print(row_data_2)
    # print(f'combined 1: ')
    # print(row_data_combo_1)
    # print(f'combined 2: ')
    # print(row_data_combo_2)
    # print(f'combined 3: ')
    # print(row_data_combo_3)

    table1 = daff.PythonTableView(row_data_combo_1) #daff needs a header in order to work correctly!
    table2 = daff.PythonTableView(row_data_combo_2)
    table3 = daff.PythonTableView(row_data_combo_3)
    
    #old version:
    # table1 = daff.PythonTableView([list(r.values()) for r in row_data_1])
    # table2 = daff.PythonTableView([list(r.values()) for r in row_data_2])

    alignment = daff.Coopy.compareTables3(table3,table2,table1).align() #3 way: initial vs server vs saving
    
    # alignment = daff.Coopy.compareTables(table3,table2).align() #initial vs server
    alignment2 = daff.Coopy.compareTables(table2,table1).align() #saving vs server
    # alignment2 = daff.Coopy.compareTables(table1, table2).align() #server vs saving
    # alignment = daff.Coopy.compareTables(table3,table1).align() #initial vs saving


    data_diff = []
    table_diff = daff.PythonTableView(data_diff)

    flags = daff.CompareFlags()

    # flags.allowDelete()
    # flags.allowUpdate()
    # flags.allowInsert()

    highlighter = daff.TableDiff(alignment2,flags)
    
    highlighter.hilite(table_diff)
    #hasDifference() should return true - and it does. 
    if highlighter.hasDifference():
        print(f'HASDIFFERENCE')
        print(highlighter.getSummary().row_deletes)
    else:
        print(f'no difference found')
    diff2html = daff.DiffRender()
    diff2html.usePrettyArrows(False)
    diff2html.render(table_diff)
    table_diff_html = diff2html.html()

    # print(table_diff_html)
    # print(f'table 1 before patch test: ')
    # print(table1.toString()) 
    # patch test: 
    # patcher = daff.HighlightPatch(table2,table_diff)
    # patcher.apply()
    # print(f'patch tester: ..................')
    # print(f'table1:')
    # print(table1.toString())
    # print(f'table2:')
    # print(table2.toString())
    # print(table2.toString()) 
    # print(f'table3:')
    # print(table3.toString()) 
    # table2String = table2.toString().strip() #no
    #todo: Task 1: turn MergeData into a Dict in order to post it to Github!
    # - use Janna's sheet builder example? 
    # - post direct instead of going through Flask front-end? 
    # table2String.strip()
    # table2Json = json.dumps(table2)
    # table2Dict = dict(todo: make this into a dict with id:0++ per row here!)
    # table2String = dict(table2String) #nope

    # merger test: 
    # print(f'Merger test: ') 
    merger = daff.Merger(table3,table2,table1,flags) #(3initial, 1saving, 2server, flags)
    merger.apply()
    # print(f'table2:')
    # table2String = table2.toString()
    # print(table2String) #after merger

    data = table2.getData() #merger table in list format
    # print(f'data: ')
    # print(json.dumps(data)) #it's a list.
    # convert to correct format (list of dicts):
    dataDict = []
    iter = -1
    for k in data:        
        # add "id" value:
        iter = iter + 1
        dictT = {}
        if iter == 0:
            pass
            # print(f'header row - not using')
        else:
            dictT['id'] = iter # add "id" with iteration
            for key, val in zip(row_header, k):      
                #deal with conflicting val?

                dictT[key] = val
        # add to list:
        if iter > 0: # not header - now empty dict
            dataDict.append( dictT ) 
        # print(f'update: ')
        # print(dataDict)

        
    
    # print(f'dataDict: ')
    # print(json.dumps(dataDict))
    # print(f'the type of dataDict is: ')
    # print(type(dataDict))

    # print(f'merger data:') #none
    # print(daff.DiffSummary().different) #nothing here? 
    # mergerConflictInfo = merger.getConflictInfos()
    
    # print(f'Merger conflict infos: ')
    # print(f'table1:')
    # print(table1.toString())
    # print(f'table2:')
    # print(table2.toString()) 
    # print(f'table3:')
    # print(table3.toString()) 
   
    return (table_diff_html, dataDict)


def searchAcrossSheets(repo_name, search_string):
    searcherAllResults = searcher.searchFor(repo_name, search_string=search_string)
    return searcherAllResults

def searchAssignedTo(repo_name, initials):
    searcherAllResults = searcher.searchFor(repo_name, assigned_user=initials)
    return searcherAllResults


if __name__ == "__main__":        # on running python app.py

    app.run(debug=app.config["DEBUG"], port=8080)        # run the flask app



# [END gae_python37_app]
